# coding=sjis

import openpyxl
import yaml
import time

from progressbar import ProgressBar
import progressbar

from entity import EntityInfo
from entity import ColumnInfo
from entity import IndexInfo
from entity import RelationInfo
from entity import Entity

"""
エンティティ定義書（エクセルファイル）を読み込んで、各エンティティのデータを格納するクラス
"""
class DefBook:
    book = []
    sheetnames = []
    entities = []

    def __init__(self):
        with open('config.yml', encoding='utf-8') as file:
            config = yaml.safe_load(file)
            self.book = openpyxl.load_workbook(config['book_name'])
            self.sheetnames = self.book.sheetnames
            del self.sheetnames[0:3]
            widgets = ['Read Worksheets     : ', progressbar.Percentage(), ' (', progressbar.Counter(), ' of ', str(len(self.sheetnames)) + ') ', progressbar.Bar(), ' ', progressbar.Timer(), ' ', progressbar.ETA(), ' ', ]
        
            pbar = ProgressBar(maxval=len(self.sheetnames), widgets=widgets).start()            
            s = 0
            for n in self.sheetnames:
                entity = Entity(self.book[n])
                self.entities.append(entity)
                s = s + 1
                time.sleep(0.1)
                pbar.update(s)
